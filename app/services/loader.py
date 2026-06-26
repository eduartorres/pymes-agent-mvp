import os
from typing import Dict, List, Optional

import openpyxl


# Ruta al Excel de solicitudes. Se resuelve relativa a este modulo.
RUTA_EXCEL = os.path.join(os.path.dirname(__file__), "../../data/solicitudes.xlsx")


def cargar_solicitudes() -> List[Dict]:
    """
    Lee el Excel de solicitudes y retorna cada fila como un diccionario.

    openpyxl en modo read_only evita cargar el libro completo en memoria,
    lo que importa cuando el archivo crece a varios miles de registros.

    Los NITs vienen en notacion cientifica desde Excel (ej. 9.0E8) porque
    la celda no tiene formato de texto. Se normalizan a string entero aqui
    para que el resto del pipeline los maneje de forma uniforme.
    """
    wb = openpyxl.load_workbook(RUTA_EXCEL, read_only=True)
    ws = wb.active

    # Primera fila como encabezados
    encabezados = [
        str(celda.value).strip() if celda.value else ""
        for celda in next(ws.iter_rows(max_row=1))
    ]

    solicitudes = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        # Filas vacias al final del libro
        if not fila[0]:
            continue

        registro = dict(zip(encabezados, fila))
        registro["nit"] = _normalizar_nit(registro.get("nit", ""))
        registro["numero_empleados_declarado"] = _normalizar_entero(
            registro.get("numero_empleados_declarado", 0)
        )
        solicitudes.append(registro)

    wb.close()
    return solicitudes


def obtener_solicitud_por_id(id_solicitud: str) -> Optional[Dict]:
    """Busca y retorna una solicitud por su ID. Retorna None si no existe."""
    for solicitud in cargar_solicitudes():
        if solicitud.get("id_solicitud") == id_solicitud:
            return solicitud
    return None


def obtener_solicitudes_rango(inicio: int = 1, fin: int = 10) -> List[Dict]:
    """
    Retorna un subconjunto de solicitudes definido por posicion (base 1).
    Util para procesar en lotes sin cargar todo en memoria de la API.
    """
    todas = cargar_solicitudes()
    return todas[inicio - 1:fin]


def _normalizar_nit(valor) -> str:
    """Convierte un NIT en formato numerico o cientifico a string entero limpio."""
    try:
        return str(int(float(str(valor))))
    except (ValueError, TypeError):
        return str(valor)


def _normalizar_entero(valor) -> int:
    """Convierte un valor numerico de Excel (puede ser float) a entero."""
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return 0
